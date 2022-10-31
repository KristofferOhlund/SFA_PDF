from PyPDF2 import PdfFileReader
import os
from openpyxl import Workbook, styles
import regex as re
import datetime


# TODO MÅNDAG: Vi loopar genom regex patterns och hittar nu för 3 olika typer
# Standard, totalvikt och nätarmering.
# När totalvikt, ska vi endast hämta från första sidan - annoteringar är fel, dvs trots 3 sidor
# kan annoteringar vara på 5 ställen. Nu printas alla filer det gäller och justering görs manuellt i Excel.
# Övriga patterns ska matcha mot samtliga sidor


def check_valid_path(path=None):
    """Läs av alla pdf i angiven sökväg"""

    src_path = None
    if path is None:
        src_path = input(r"Välj sökväg för att läsa pdf")
    else:
        src_path = path

    return src_path


def read_from_path(path, flatten=False):
    """Läser alla PDF som finns i path. 
    Om flatten == False, specar varje sida och vikt för sig.
    Om flatten == True, summerar totalvikten för hela filen istället för vikt på resp sida """
    
    # Data
    file_data = {}

    # Specar vikt per sida
    if flatten == False:
        for file in os.listdir(path):
            if file.endswith(".pdf"):
                print(file)
                pdf_with_path = os.path.join(path, file)
                # läs pdf
                pdf_reader = PdfFileReader(pdf_with_path)
                pages = pdf_reader.getNumPages()
                # Regex, vikt kommer via splitlies tillsammans med armeringsförteckning.
                # pattern nedan matchar PDF som har rad STANDARDFÖRTECKNING 0 Vikt
                vikt_pattern = r"(\d+)\sARMERINGSFÖRTECKNING" 
                # Om PDF har text Totalvikt - kg 0 vikt, använd nedan regex
                #vikt_pattern = r"TOTAL VIKT kg (\d+)"

                # läs varje page
                page_data = {}
                for page in range(0, pages):
                    text = pdf_reader.getPage(page).extract_text().splitlines()
                    # Data för page[i]
                    for line in text:
                        vikt = re.search(vikt_pattern, line)
                        if vikt is not None:
                            page_data[f"SIDA {page + 1}"] = page + 1
                            page_data[f"Vikt sida {page + 1}"] = vikt.group(1)
                
                # lägg till fil + pagedata i huvud data
                file_data[file] = page_data
    
    # Summerar vikt per fil
    else:
        for file in os.listdir(path):
            läs_en_gång = ""
            if file.endswith(".pdf"):
                pdf_with_path = os.path.join(path, file)
                # läs pdf
                pdf_reader = PdfFileReader(pdf_with_path)
                pages = pdf_reader.getNumPages()
                # Regex, vikt kommer via splitlies tillsammans med armeringsförteckning.
                # pattern nedan matchar PDF som har rad STANDARDFÖRTECKNING 0 Vikt
               # vikt_pattern = r"(\d+)\sARMERINGSFÖRTECKNING"
                # Om PDF har text Totalvikt - kg 0 vikt, använd nedan regex
                vikt_pattern = r"TOTAL VIKT kg (\d+)ARMERINGSFÖRTECKNING"

                # REGEX PATTERNS 
                patterns = [r"TOTAL VIKT kg (\d+)ARMERINGSFÖRTECKNING",
                            r"(\d+)\sARMERINGSFÖRTECKNING", r"(\d+)\sNÄTFÖRTECKNING"]

                # vikt per fil
                vikt_per_fil = 0
                # läs varje page
                for page in range(0, pages):
                    text = pdf_reader.getPage(page).extract_text().splitlines()
                    # Testar alla Regex på varje line
                    for pattern in patterns:
                        # Data för page[i]
                        for line in text:
                            vikt = re.search(pattern, line)
                            if vikt is not None and pattern == patterns[0]:
                                vikt_per_fil += int(vikt.group(1))
                                läs_en_gång = file
                            elif vikt is not None and pattern == patterns[1]:
                                vikt_per_fil += int(vikt.group(1))
                            elif vikt is not None and pattern == patterns[2]:
                                vikt_per_fil += int(vikt.group(1))
                
                # Justera följande i excel
                if läs_en_gång != "":
                    print(f"FÖLJANDE SKA LÄSAS EN GÅNG {läs_en_gång}")

                # lägg till fil + pagedata i huvud data
                file_data[file] = vikt_per_fil
    

    return file_data



def write_excel_summary(data: dict, output, flatten=False):
    """Skriv till excel, spara i output.
    Om flatten == False, skriver varje sida på ny rad. Kolumner = fil, sida, vikt
    Om flatten == True, skriver bara fil och totalvikt. Kolumner = fil, vikt. """
    
    wb = Workbook()
    ws = wb.active

    bold = styles.Font(bold=True)

    if flatten == False:
        # Skriver headers
        headers = ["NAMN", "SIDA", "VIKT - KG"]
        header_row = 1
        header_col = 1

        for header in headers:
            ws.cell(header_row, header_col, value=header).font = bold
            header_col += 1

        # skriver data
        data_row = 3
        data_col = 1
        start_data_col = 1
        multiple_pages_col_start = 2

        for k in data:
            ws.cell(data_row, data_col).value = k
            data_col += 1
            for value in data[k]:
                ws.cell(data_row, data_col).value = int(data[k][value])
                if data_col < 3:
                    data_col += 1
                else:
                    data_row += 1
                    data_col = multiple_pages_col_start
            
            # Börja om med nästa fil
            data_col = start_data_col

    # Data är flattend, skriver bara fil och totalvikt
    else:
        # Skriver headers
        headers = ["NAMN", "TOTALVIKT - KG"]
        header_row = 1
        header_col = 1

        for header in headers:
            ws.cell(header_row, header_col, value=header).font = bold
            header_col += 1

        # Skriver data
        data_row = 3
        data_col = 1
        data_reset_col = 1
        for key in data:
            ws.cell(data_row, data_col).value = key
            data_col += 1
            ws.cell(data_row, data_col).value = data[key]
            data_row += 1
            data_col = data_reset_col

    avläst_dir = os.path.basename(output)
    excel = os.path.join(output, f"Totalvikt_Armering_{avläst_dir}.xlsx")
    wb.save(excel)

    print(f"Antal avlästa filer: {len(data)}")




if __name__ == "__main__":
    # Välj sökväg för mapp att skanna av, excel rapport hamnar i samma mapp som du skannat.
    # Om flatten, hämtar totalvikt per fil, annars räknar vikt per sida 
    # Om flatten, skriver endast 2 kolumner, fil och total vikt. Annars 3 kolumner, fil, sida, vikt
    p = input(r"Välj sökväg till mapp du vill skanna av: ")
    valid_path = check_valid_path(p)
    data = read_from_path(valid_path, flatten=True)
    to_excel = write_excel_summary(data, output=p, flatten=True)
    print("GÖR STICKPROV 10-20 RITNINGAR FRÅN VARJE SÖKVÄG")
